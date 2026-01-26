import os
import time     
import pandas                   # standard for managing tables in 
from datetime import datetime   # timestamp for each test
from config import EXCEL_FILE, Colors as C

# for better styling excel file
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# create the dictionary (factory pattern)
def init_run_data(model_name, project_name, prompt):
    return {
        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Model": model_name,
        "Project_Name": project_name,
        "Prompt": prompt,
        "Status": "PENDING",
        "Gen_Time_s": 0,
        "Exec_Time_s": 0,
        "Volume_mm3": 0,
        "Faces_Count": 0,
        "Error_Log": "",
        "Code_Lines": 0,
        "Library": "None"
    }

# to adjust column width and row coloring based on status (SUCCESS, FAIL...)
def format_excel_file(filepath):
    try:
        wb = load_workbook(filepath)    # loading excel file as an editable object
        ws = wb.active                  # considering the active sheet

        # defining colors (HEX codes)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Verde pastello
        green_font = Font(color="006100") 
        
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Rosso pastello
        red_font = Font(color="9C0006")   

        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Giallo
        yellow_font = Font(color="9C6500")

        # adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column) # 'A', 'B', 'C', ...
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
           
            adjusted_width = (max_length + 2)   # setting width with a little margin
            if adjusted_width > 35:             # setting a limit to avoid infinite column
                adjusted_width = 35
            
            ws.column_dimensions[column_letter].width = adjusted_width

        # searching for status column
        status_col_index = None
        for cell in ws[1]: 
            if cell.value == "Status":
                status_col_index = cell.column
                break

        # conditional coloring status based
        if status_col_index:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=status_col_index)
                status_text = str(cell.value).upper()

                if "SUCCESS" in status_text:
                    cell.fill = green_fill
                    cell.font = green_font
                elif "ERROR" in status_text or "FAIL" in status_text:
                    cell.fill = red_fill
                    cell.font = red_font
                elif "EMPTY" in status_text or "PENDING" in status_text:
                    cell.fill = yellow_fill
                    cell.font = yellow_font

        wb.save(filepath)
        print("Excel formatted successfully.")

    except Exception as e:
        print(f"{C.RED}ERROR: Unable to format Excel.\n{e}{C.END}")

# function to manage excel file (creation/reading/writing/errors)
def save_to_excel(obj_data):

    # pass the object data dictionary into a list to prevent pandas from getting confused
    new_df = pandas.DataFrame([obj_data])

    # save file retry logic
    max_retries = 5
    for attempt in range(max_retries):
        try:    
            if os.path.exists(EXCEL_FILE):
                # read the entire file
                existing_df = pandas.read_excel(EXCEL_FILE)
                # create a history of changes, discarding old row/column index
                final_df = pandas.concat([existing_df, new_df], ignore_index = True)
            else:
                final_df = new_df 
        
            # write the file to disk without the row number column
            final_df.to_excel(EXCEL_FILE, index = False)
            format_excel_file(EXCEL_FILE) 
            print(f"Excel file updated successfully: {EXCEL_FILE}")
            return
        
        # if we try to edit the excel file while it's open, an error will occur
        except PermissionError:
            print(f"""{C.YELLOW}
            WARNING: the excel file '{EXCEL_FILE}' seems to be open!\n
            Please close it within 5 second, otherwise all data will be lost.\n
            Attempt {attempt+1}/{max_retries}.{C.END}
            """)

        except Exception as e:
            print(f"{C.RED}ERROR: an unexpected error occurred while saving file.\n{e}{C.END}")
            break

    print(f"{C.RED}ERROR: UNABLE TO SAVE TO EXCEL. The data from this run is lost or only printed to screen.{C.END}")